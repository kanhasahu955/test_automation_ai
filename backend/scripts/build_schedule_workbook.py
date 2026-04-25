"""Generate the **schedule execution** Excel workbook.

A two-purpose workbook:

  1. **Bulk import template** — fill the `Schedules` sheet with one row per
     scheduled execution (target type/id, cadence, params), then submit it
     via the upcoming bulk-schedule import. The shape matches the
     `RunType` enum and the existing execution endpoints, so a future
     bulk-import controller can read this file directly.

  2. **Live reference** — sheets for cron syntax, timezone codes, retry
     policy fields, and the existing executions API surface, so anyone
     scheduling work has a single artefact to fill in + read.

Sheets:
  * Instructions
  * Schedules            (template + drop-downs)
  * Examples             (one row per common cadence + target combo)
  * Cron Reference       (cron expression cookbook)
  * Run Types            (enum reference)
  * Run Status           (enum reference)
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from app.modules.executions.models import ResultStatus, RunStatus, RunType
from scripts._xlsx_style import (
    autosize,
    configure_workbook,
    section_title,
    style_hero,
    write_kv_block,
    write_table,
)

TARGET_TYPES = ["test_case", "test_suite", "no_code_flow", "stm_document"]


def _build_instructions(wb: Workbook) -> None:
    ws = wb.active
    assert ws is not None
    ws.title = "Instructions"

    style_hero(
        ws,
        title="Schedule Execution — QualityForge AI",
        subtitle="Plan, document and bulk-import scheduled test runs.",
        columns=2,
    )

    write_kv_block(
        ws,
        rows=[
            ("Audience", "QA leads, ops engineers, release managers."),
            ("Purpose", "Capture every recurring or one-off scheduled execution in a single file."),
            ("How to use", "Edit the Schedules sheet → review with your team → import."),
            ("Run types", ", ".join(rt.value for rt in RunType)),
            ("Default timezone", "UTC. Use IANA names like 'America/New_York' if you need otherwise."),
            ("Cron format", "Standard 5-field crontab — minute hour day-of-month month day-of-week"),
            ("Live API", "GET /api/v1/projects/{project_id}/execution-runs"),
            ("Live API", "GET /api/v1/execution-runs/{run_id}/report"),
        ],
        start_row=4,
    )

    section_title(ws, row=14, text="Column reference (Schedules sheet)", span=2)

    rows: list[list[object]] = [
        ["schedule_name", "Human label, e.g. 'Nightly regression — checkout'.", "yes"],
        ["target_type", "What is being executed: test_case / test_suite / no_code_flow / stm_document.", "yes"],
        ["target_id", "UUID of the target entity (suite, flow, etc.).", "yes"],
        ["project_id", "Project that owns the target.", "yes"],
        ["environment", "dev / staging / production / custom env name.", "no"],
        ["run_type", "MANUAL, SCHEDULED, CI_CD or AIRFLOW. Use SCHEDULED for cron entries.", "yes"],
        ["cron_expression", "5-field cron, e.g. '0 2 * * *' for daily at 02:00.", "yes (if SCHEDULED)"],
        ["timezone", "IANA TZ name (default UTC).", "no"],
        ["start_at", "ISO 8601 timestamp — when the schedule becomes active.", "no"],
        ["end_at", "ISO 8601 timestamp — when the schedule expires (optional).", "no"],
        ["max_retries", "Integer ≥ 0. Worker retries on transient failures.", "no"],
        ["retry_backoff_s", "Integer seconds between retries.", "no"],
        ["timeout_s", "Hard timeout per run. Defaults to suite timeout.", "no"],
        ["data_source_id", "Specific data source to bind for STM/profiling runs.", "no"],
        ["params_json", "JSON payload merged into the run context.", "no"],
        ["notify_emails", "Comma-separated emails to notify on completion.", "no"],
        ["notify_slack", "Slack webhook URL (or @channel-name with bot installed).", "no"],
        ["enabled", "TRUE/FALSE. Disable without deleting.", "yes"],
        ["owner", "Person responsible (email or name).", "yes"],
        ["notes", "Free-text justification / change-control reference.", "no"],
    ]
    write_table(
        ws,
        headers=["Column", "Description", "Required"],
        rows=rows,
        start_row=15,
        column_widths=[22, 80, 16],
        table_name="ScheduleColumnRef",
    )


SCHEDULE_COLUMNS = [
    "schedule_name", "target_type", "target_id", "project_id", "environment",
    "run_type", "cron_expression", "timezone", "start_at", "end_at",
    "max_retries", "retry_backoff_s", "timeout_s", "data_source_id",
    "params_json", "notify_emails", "notify_slack", "enabled", "owner", "notes",
]


def _build_schedules(wb: Workbook) -> None:
    ws = wb.create_sheet("Schedules")
    style_hero(
        ws,
        title="Schedules — fill rows below",
        subtitle="Drop-downs are enforced on Run Type, Target Type and Enabled.",
        columns=len(SCHEDULE_COLUMNS),
    )

    write_table(
        ws,
        headers=SCHEDULE_COLUMNS,
        rows=[],
        start_row=4,
        column_widths=[
            28, 16, 38, 38, 14, 14, 18, 18, 24, 24,
            12, 18, 14, 38, 40, 28, 30, 10, 24, 50,
        ],
        table_name="SchedulesTemplate",
    )

    # ---- Drop-down: target_type ----
    target_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(TARGET_TYPES)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid target type",
        error="Use one of: " + ", ".join(TARGET_TYPES),
    )
    ws.add_data_validation(target_dv)
    target_col = chr(ord("A") + SCHEDULE_COLUMNS.index("target_type"))
    target_dv.add(f"{target_col}5:{target_col}204")

    # ---- Drop-down: run_type ----
    run_type_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(rt.value for rt in RunType)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid run type",
        error="Use one of: " + ", ".join(rt.value for rt in RunType),
    )
    ws.add_data_validation(run_type_dv)
    rt_col = chr(ord("A") + SCHEDULE_COLUMNS.index("run_type"))
    run_type_dv.add(f"{rt_col}5:{rt_col}204")

    # ---- Drop-down: enabled ----
    enabled_dv = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="TRUE or FALSE only",
    )
    ws.add_data_validation(enabled_dv)
    en_col = chr(ord("A") + SCHEDULE_COLUMNS.index("enabled"))
    enabled_dv.add(f"{en_col}5:{en_col}204")


def _build_examples(wb: Workbook) -> None:
    ws = wb.create_sheet("Examples")
    style_hero(
        ws,
        title="Schedule examples — copy a row to the Schedules sheet",
        subtitle="Realistic cadences for the most common QualityForge use-cases.",
        columns=len(SCHEDULE_COLUMNS),
    )

    rows: list[list[object]] = [
        # 1. Nightly full regression
        [
            "Nightly regression — full suite",
            "test_suite",
            "11111111-1111-1111-1111-111111111111",
            "00000000-0000-0000-0000-000000000001",
            "production",
            "SCHEDULED",
            "0 2 * * *",
            "UTC",
            "2026-05-01T00:00:00Z",
            "",
            2, 60, 1800, "",
            '{"browser":"chrome","headless":true}',
            "qa-leads@example.com",
            "https://hooks.slack.com/services/T0/B0/XYZ",
            "TRUE",
            "qa-lead@example.com",
            "Runs every night before US trading open.",
        ],
        # 2. Hourly smoke flow
        [
            "Hourly smoke — checkout flow",
            "no_code_flow",
            "22222222-2222-2222-2222-222222222222",
            "00000000-0000-0000-0000-000000000001",
            "staging",
            "SCHEDULED",
            "0 * * * *",
            "UTC",
            "", "",
            1, 30, 600, "",
            '{"region":"eu-west-1"}',
            "ops@example.com",
            "",
            "TRUE",
            "ops@example.com",
            "Smoke check every hour during business days.",
        ],
        # 3. Weekly STM validation
        [
            "Weekly STM — customer 360",
            "stm_document",
            "33333333-3333-3333-3333-333333333333",
            "00000000-0000-0000-0000-000000000002",
            "production",
            "AIRFLOW",
            "0 4 * * 1",
            "America/New_York",
            "", "",
            0, 0, 7200,
            "44444444-4444-4444-4444-444444444444",
            '{"sample_size":100000,"strict":true}',
            "data-quality@example.com",
            "@data-quality",
            "TRUE",
            "data-quality@example.com",
            "Run after the Sunday warehouse refresh.",
        ],
        # 4. CI-driven (no cron, fired by the CI pipeline)
        [
            "PR pipeline — quick suite",
            "test_suite",
            "55555555-5555-5555-5555-555555555555",
            "00000000-0000-0000-0000-000000000001",
            "ephemeral",
            "CI_CD",
            "",
            "UTC",
            "", "",
            0, 0, 900, "",
            '{"trigger":"github-actions"}',
            "",
            "",
            "TRUE",
            "platform@example.com",
            "Triggered from GitHub Actions on PR open.",
        ],
        # 5. Manual sanity check (placeholder)
        [
            "Pre-release manual sanity",
            "test_case",
            "66666666-6666-6666-6666-666666666666",
            "00000000-0000-0000-0000-000000000001",
            "staging",
            "MANUAL",
            "",
            "UTC",
            "", "",
            0, 0, 0, "",
            "",
            "release-eng@example.com",
            "",
            "FALSE",
            "release-eng@example.com",
            "Disabled by default; flipped to TRUE only during release windows.",
        ],
        # 6. Quarter-hour data profiling on staging
        [
            "Profiling — staging customers (15-min)",
            "stm_document",
            "77777777-7777-7777-7777-777777777777",
            "00000000-0000-0000-0000-000000000003",
            "staging",
            "SCHEDULED",
            "*/15 * * * *",
            "UTC",
            "", "",
            3, 15, 300,
            "88888888-8888-8888-8888-888888888888",
            '{"profile_type":"drift","baseline_id":"99999999"}',
            "data-quality@example.com",
            "",
            "TRUE",
            "data-quality@example.com",
            "Drift watcher; pages on three consecutive failures.",
        ],
    ]
    write_table(
        ws,
        headers=SCHEDULE_COLUMNS,
        rows=rows,
        start_row=4,
        column_widths=[
            36, 16, 38, 38, 14, 14, 18, 18, 24, 24,
            12, 18, 14, 38, 40, 28, 30, 10, 24, 50,
        ],
        table_name="ScheduleExamples",
    )


def _build_cron_reference(wb: Workbook) -> None:
    ws = wb.create_sheet("Cron Reference")
    style_hero(
        ws,
        title="Cron expressions — cookbook",
        subtitle="Standard 5-field crontab. Times are interpreted in the schedule's timezone.",
        columns=3,
    )
    write_table(
        ws,
        headers=["Cadence", "Cron expression", "Notes"],
        rows=[
            ["Every minute", "* * * * *", "Avoid for production — use Celery beat or in-process timers instead."],
            ["Every 5 minutes", "*/5 * * * *", "Useful for liveness probes."],
            ["Every 15 minutes", "*/15 * * * *", ""],
            ["Every hour (top of hour)", "0 * * * *", ""],
            ["Every 4 hours", "0 */4 * * *", ""],
            ["Daily at 02:00", "0 2 * * *", "Recommended for nightly regression."],
            ["Daily at 02:30 UTC", "30 2 * * *", "Pair with timezone=UTC."],
            ["Twice daily", "0 6,18 * * *", "Comma-separated list."],
            ["Weekdays only at 09:00", "0 9 * * 1-5", "1=Mon, 5=Fri."],
            ["Weekly Monday 04:00", "0 4 * * 1", ""],
            ["Monthly on the 1st at 03:00", "0 3 1 * *", ""],
            ["Quarterly (Jan/Apr/Jul/Oct)", "0 3 1 1,4,7,10 *", ""],
            ["First Monday of month at 09:00", "0 9 1-7 * 1", "Day-of-month and day-of-week combined (cron AND)."],
            ["Last day of month — best-effort", "0 23 28-31 * *", "Use Airflow `@monthly` for true last-day semantics."],
            ["Special: @hourly", "@hourly", "Some implementations support shorthand."],
            ["Special: @daily", "@daily", "Equivalent to '0 0 * * *'."],
            ["Special: @weekly", "@weekly", "Equivalent to '0 0 * * 0'."],
            ["Special: @monthly", "@monthly", "Equivalent to '0 0 1 * *'."],
        ],
        start_row=4,
        column_widths=[36, 22, 70],
        table_name="CronReference",
    )

    section_title(ws, row=26, text="Field reference", span=3)
    write_table(
        ws,
        headers=["Field", "Range", "Notes"],
        rows=[
            ["minute", "0-59", ""],
            ["hour", "0-23", "0 = midnight in the schedule's timezone."],
            ["day-of-month", "1-31", ""],
            ["month", "1-12", "Or JAN-DEC."],
            ["day-of-week", "0-6", "0 = Sunday. Or SUN-SAT."],
        ],
        start_row=27,
        column_widths=[18, 12, 60],
    )


def _build_enum_sheet(wb: Workbook, *, name: str, title: str, subtitle: str, enum_cls) -> None:
    ws = wb.create_sheet(name)
    style_hero(ws, title=title, subtitle=subtitle, columns=3)
    rows: list[list[object]] = []
    for member in enum_cls:
        rows.append([member.value, member.name, _enum_help(name, member.value)])
    write_table(
        ws,
        headers=["Value", "Name", "When to use"],
        rows=rows,
        start_row=4,
        column_widths=[18, 18, 80],
    )
    autosize(ws, headers=["A", "B", "C"], max_width=80)


_ENUM_HELP: dict[tuple[str, str], str] = {
    ("Run Types", "MANUAL"): "Triggered ad-hoc by a user from the UI or API.",
    ("Run Types", "SCHEDULED"): "Started by Celery beat / cron based on cron_expression.",
    ("Run Types", "CI_CD"): "Started by an external CI pipeline (GitHub Actions, GitLab CI, …).",
    ("Run Types", "AIRFLOW"): "Triggered by an Airflow DAG task.",
    ("Run Status", "PENDING"): "Run is queued — no worker has picked it up yet.",
    ("Run Status", "RUNNING"): "Worker is actively executing the run.",
    ("Run Status", "PASSED"): "Every test in the run passed.",
    ("Run Status", "FAILED"): "At least one test in the run failed.",
    ("Run Status", "CANCELLED"): "Run was cancelled — by user, timeout, or system.",
    ("Result Status", "PASSED"): "Test produced expected output.",
    ("Result Status", "FAILED"): "Assertion failed — see error_message + result_json for details.",
    ("Result Status", "SKIPPED"): "Test was intentionally skipped (precondition not met).",
    ("Result Status", "ERROR"): "Test errored before assertions ran (setup/teardown issue).",
}


def _enum_help(sheet_name: str, value: str) -> str:
    return _ENUM_HELP.get((sheet_name, value), "")


def build(output: Path) -> Path:
    wb = Workbook()
    _build_instructions(wb)
    _build_schedules(wb)
    _build_examples(wb)
    _build_cron_reference(wb)
    _build_enum_sheet(
        wb,
        name="Run Types",
        title="Run Types",
        subtitle="What kicked the execution off.",
        enum_cls=RunType,
    )
    _build_enum_sheet(
        wb,
        name="Run Status",
        title="Run Status",
        subtitle="Lifecycle of an execution_run row.",
        enum_cls=RunStatus,
    )
    _build_enum_sheet(
        wb,
        name="Result Status",
        title="Result Status",
        subtitle="Per-test outcome inside a run.",
        enum_cls=ResultStatus,
    )
    configure_workbook(wb)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "docs" / "excel" / "qualityforge-ai_schedule_executions.xlsx",
    )
    args = parser.parse_args()
    path = build(args.out)
    print(f"✓ wrote {path}")


if __name__ == "__main__":
    main()
