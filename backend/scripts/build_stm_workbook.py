"""Generate the **STM upload template** Excel workbook.

This is the file users download, fill in their Source-to-Target mappings, and
upload via `POST /projects/{project_id}/stm/upload`. The backend's
`parse_stm_excel` accepts these exact column aliases — see
`app/utils/file_parser.py`.

Sheets produced:
  * **Instructions** — what each column means + how the upload flow works.
  * **Mappings** — empty template with header row + format the parser accepts.
  * **Examples** — end-to-end scenarios for each `ValidationType` so users can
    copy a row and adapt it.
  * **Lookups** — enumerations (validation types, tips, common pitfalls).

The mapping sheet uses Excel data validation drop-downs for `validation_type`
so users cannot enter an invalid value.
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from app.modules.stm_converter.models import ValidationType
from app.utils.file_parser import EXPECTED_COLUMNS
from scripts._xlsx_style import (
    autosize,
    configure_workbook,
    section_title,
    style_hero,
    write_kv_block,
    write_table,
)

# Canonical column names the parser uses (first alias for each canonical key).
COLUMNS: list[str] = [
    "source_table",
    "source_column",
    "target_table",
    "target_column",
    "join_key",
    "transformation_rule",
    "validation_type",
]
COLUMN_HEADERS = {
    "source_table": "Source Table",
    "source_column": "Source Column",
    "target_table": "Target Table",
    "target_column": "Target Column",
    "join_key": "Join Key",
    "transformation_rule": "Transformation Rule",
    "validation_type": "Validation Type",
}


def _build_instructions(wb: Workbook) -> None:
    ws = wb.active
    assert ws is not None
    ws.title = "Instructions"
    style_hero(
        ws,
        title="STM Upload Template — QualityForge AI",
        subtitle="Fill the Mappings sheet and upload via the STM Converter (or POST /api/v1/projects/{id}/stm/upload).",
        columns=2,
    )

    write_kv_block(
        ws,
        rows=[
            ("Endpoint", "POST /api/v1/projects/{project_id}/stm/upload"),
            ("Auth", "Bearer JWT (must have data-write permission)"),
            ("Form field", "file=<this workbook>.xlsx"),
            ("Accepted extensions", ".xlsx, .xls"),
            ("First row", "Header row — must contain the columns listed below."),
            ("Header aliases", "The parser is forgiving: 'src_table', 'source table' etc. all work."),
            ("Empty rows", "Rows with no source_table AND no target_table are skipped."),
            ("After upload", "Document is parsed → mappings stored → SQL can be generated."),
        ],
        start_row=4,
    )

    section_title(ws, row=14, text="Column reference", span=2)

    column_rows: list[list[object]] = []
    descriptions = {
        "source_table": "Fully-qualified source table (schema.table or just table).",
        "source_column": "Source column name. Required for column-level checks.",
        "target_table": "Target table where the data lands after transformation.",
        "target_column": "Target column to validate against.",
        "join_key": "Column(s) used to join source ↔ target. Comma-separate for composites.",
        "transformation_rule": "SQL-friendly expression on source columns (e.g. UPPER(first_name)).",
        "validation_type": "One of: " + ", ".join(v.value for v in ValidationType),
    }
    required = {"source_table", "target_table", "validation_type"}
    for canonical in COLUMNS:
        column_rows.append(
            [
                COLUMN_HEADERS[canonical],
                "yes" if canonical in required else "no",
                ", ".join(EXPECTED_COLUMNS[canonical]),
                descriptions[canonical],
            ]
        )

    write_table(
        ws,
        headers=["Column", "Required", "Accepted aliases", "Description"],
        rows=column_rows,
        start_row=15,
        column_widths=[22, 12, 38, 80],
        table_name="STMColumnReference",
    )

    section_title(ws, row=24, text="Validation types — what each one does", span=2)
    write_table(
        ws,
        headers=["Validation Type", "What it checks", "Required columns beyond source/target"],
        rows=[
            ["ROW_COUNT", "Source row count == target row count.", "source_table, target_table"],
            ["NULL_CHECK", "Target column has no NULL values for matching rows.", "target_column"],
            ["DUPLICATE_CHECK", "No duplicate values in the target column (within the join scope).", "target_column, join_key"],
            ["TRANSFORMATION_CHECK", "Applying transformation_rule to source produces target.", "source_column, target_column, transformation_rule, join_key"],
            ["REFERENCE_CHECK", "Every target row has a matching row in source via join_key.", "join_key"],
        ],
        start_row=25,
        column_widths=[24, 60, 50],
    )


def _build_mappings(wb: Workbook) -> None:
    ws = wb.create_sheet("Mappings")
    style_hero(
        ws,
        title="Mappings — fill rows below and upload",
        subtitle="Header row is fixed. The parser accepts any of the documented aliases.",
        columns=len(COLUMNS),
    )

    headers = [COLUMN_HEADERS[c] for c in COLUMNS]
    write_table(
        ws,
        headers=headers,
        rows=[],
        start_row=4,
        column_widths=[24, 22, 24, 22, 18, 60, 22],
        table_name="STMMappingsTemplate",
    )

    # Reserve 200 rows for entry, with a drop-down on validation_type.
    validation_options = ",".join(v.value for v in ValidationType)
    dv = DataValidation(
        type="list",
        formula1=f'"{validation_options}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Pick one of: " + validation_options,
    )
    ws.add_data_validation(dv)
    last_col_letter = chr(ord("A") + COLUMNS.index("validation_type"))
    dv.add(f"{last_col_letter}5:{last_col_letter}204")


def _build_examples(wb: Workbook) -> None:
    ws = wb.create_sheet("Examples")
    style_hero(
        ws,
        title="Example scenarios — copy a row to Mappings",
        subtitle="One row per validation type plus a few real-world variants.",
        columns=len(COLUMNS) + 1,
    )

    rows: list[list[object]] = [
        # ---- ROW_COUNT ----
        [
            "Row counts match between staging and dim_customer",
            "stg.customers", "", "dim.customer", "",
            "customer_id", "", "ROW_COUNT",
        ],
        # ---- NULL_CHECK ----
        [
            "Email column must never be null in dim_customer",
            "stg.customers", "email", "dim.customer", "email",
            "customer_id", "", "NULL_CHECK",
        ],
        [
            "Created_at must never be null in fact_orders",
            "stg.orders", "created_at", "fact.orders", "created_at",
            "order_id", "", "NULL_CHECK",
        ],
        # ---- DUPLICATE_CHECK ----
        [
            "Customer email is unique per row in dim_customer",
            "stg.customers", "email", "dim.customer", "email",
            "customer_id", "", "DUPLICATE_CHECK",
        ],
        # ---- TRANSFORMATION_CHECK ----
        [
            "Customer name uppercased on the way to dim",
            "stg.customers", "first_name", "dim.customer", "first_name_upper",
            "customer_id", "UPPER(first_name)", "TRANSFORMATION_CHECK",
        ],
        [
            "Order amount converted to USD",
            "stg.orders", "amount_local", "fact.orders", "amount_usd",
            "order_id", "amount_local * fx_rate", "TRANSFORMATION_CHECK",
        ],
        # ---- REFERENCE_CHECK ----
        [
            "Every order references an existing customer",
            "dim.customer", "", "fact.orders", "",
            "customer_id", "", "REFERENCE_CHECK",
        ],
        [
            "Every order line item references an existing order",
            "fact.orders", "", "fact.order_items", "",
            "order_id", "", "REFERENCE_CHECK",
        ],
        # ---- Composite join key ----
        [
            "Composite join — date + branch must reference dim_branch_calendar",
            "dim.branch_calendar", "", "fact.daily_sales", "",
            "branch_id, sale_date", "", "REFERENCE_CHECK",
        ],
    ]

    write_table(
        ws,
        headers=["Scenario", *[COLUMN_HEADERS[c] for c in COLUMNS]],
        rows=rows,
        start_row=4,
        column_widths=[60, 24, 22, 24, 22, 22, 50, 22],
        table_name="STMScenarios",
    )


def _build_lookups(wb: Workbook) -> None:
    ws = wb.create_sheet("Lookups")
    style_hero(
        ws,
        title="Lookups",
        subtitle="Reference data — copy values from here when filling Mappings.",
        columns=3,
    )
    write_table(
        ws,
        headers=["Validation Type", "Description", "Generated SQL template"],
        rows=[
            ["ROW_COUNT", "Source vs target count.", "templates/stm/row_count.sql"],
            ["NULL_CHECK", "Target column NULL count.", "templates/stm/null_check.sql"],
            ["DUPLICATE_CHECK", "Duplicate detection.", "templates/stm/duplicate_check.sql"],
            ["TRANSFORMATION_CHECK", "Compare transformed source vs target.", "templates/stm/transformation_check.sql"],
            ["REFERENCE_CHECK", "Referential integrity check.", "templates/stm/reference_check.sql"],
        ],
        start_row=4,
        column_widths=[24, 50, 38],
    )

    section_title(ws, row=12, text="Common pitfalls", span=3)
    write_table(
        ws,
        headers=["Pitfall", "Why it happens", "Fix"],
        rows=[
            [
                "Header not detected",
                "Column name doesn't match any alias.",
                "Use one of the aliases in the Instructions sheet.",
            ],
            [
                "Validation type silently downgraded",
                "Misspelled enum value, e.g. NULL_CHK.",
                "Use the drop-down on Mappings → Validation Type.",
            ],
            [
                "Composite join key only checks first column",
                "Using a single column where the FK is composite.",
                "Comma-separate the columns: 'branch_id, sale_date'.",
            ],
            [
                "Transformation SQL uses unqualified columns",
                "Renderer expects raw source column names.",
                "Write `UPPER(first_name)` — the renderer adds the `s.` prefix.",
            ],
            [
                "Empty rows blocking parser",
                "Non-mapping comment rows in the middle of the sheet.",
                "Either delete them or leave both source_table and target_table empty (they're skipped).",
            ],
        ],
        start_row=13,
        column_widths=[34, 50, 60],
    )

    autosize(ws, headers=["A", "B", "C"], max_width=80)


def build(output: Path) -> Path:
    wb = Workbook()
    _build_instructions(wb)
    _build_mappings(wb)
    _build_examples(wb)
    _build_lookups(wb)
    configure_workbook(wb)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "docs" / "excel" / "qualityforge-ai_stm_upload_template.xlsx",
    )
    args = parser.parse_args()
    path = build(args.out)
    print(f"✓ wrote {path}")


if __name__ == "__main__":
    main()
