"""Generate the QualityForge AI **complete API reference** Excel workbook.

The script reads the live OpenAPI spec produced by the FastAPI app (no server
required — it instantiates the app in-process) and turns it into a multi-sheet
workbook:

  * **Index** — every endpoint at a glance, hyperlinked to its detail sheet.
  * **<Tag>** — one sheet per API group (auth, projects, executions, …) with
    the request/response shape, parameters, status codes, security and a
    JSON example.
  * **Schemas** — one row per Pydantic model with all fields, types,
    nullability, defaults and descriptions.
  * **Enums** — every enumeration value the API exposes (RunStatus,
    ValidationType, …) so the spreadsheet doubles as a lookup reference.

The output guarantees the spreadsheet stays in sync with the codebase:
when a controller changes, regenerate and re-ship.
"""
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from scripts._xlsx_style import (
    autosize,
    configure_workbook,
    section_title,
    style_hero,
    write_kv_block,
    write_table,
)

# ---------------------------------------------------------------------------
# OpenAPI extraction
# ---------------------------------------------------------------------------


def _load_openapi() -> dict[str, Any]:
    """Build the FastAPI app and return its OpenAPI document."""
    # Imported lazily so that running this script never starts the DB engine —
    # we just need the route schema, which FastAPI computes statically.
    from app.main import app

    return app.openapi()


def _resolve_ref(spec: dict[str, Any], ref: str) -> dict[str, Any]:
    """Resolve a `$ref` like `#/components/schemas/UserRead` to its definition."""
    parts = ref.lstrip("#/").split("/")
    node: Any = spec
    for part in parts:
        node = node.get(part, {}) if isinstance(node, dict) else {}
    return node if isinstance(node, dict) else {}


def _schema_summary(schema: dict[str, Any], spec: dict[str, Any]) -> str:
    """Return a human-friendly one-liner for a JSON-Schema node."""
    if not schema:
        return ""
    if "$ref" in schema:
        ref = schema["$ref"]
        target = ref.split("/")[-1]
        return f"{target} (object)"
    if "anyOf" in schema or "oneOf" in schema:
        variants = schema.get("anyOf") or schema.get("oneOf") or []
        labels = [_schema_summary(v, spec) for v in variants if v]
        return " | ".join(filter(None, labels)) or "any"
    if "allOf" in schema:
        labels = [_schema_summary(v, spec) for v in schema["allOf"]]
        return " & ".join(filter(None, labels)) or "object"
    t = schema.get("type")
    if t == "array":
        item = _schema_summary(schema.get("items") or {}, spec)
        return f"array<{item or 'any'}>"
    if t == "object":
        return "object"
    fmt = schema.get("format")
    if fmt:
        return f"{t}<{fmt}>"
    if isinstance(t, list):
        return " | ".join(t)
    return t or "any"


def _example_for_schema(schema: dict[str, Any], spec: dict[str, Any], depth: int = 0) -> Any:
    """Build a best-effort JSON example for a schema (recursive, depth-capped)."""
    if depth > 4 or not schema:
        return None
    if "$ref" in schema:
        return _example_for_schema(_resolve_ref(spec, schema["$ref"]), spec, depth + 1)
    if "example" in schema:
        return schema["example"]
    if "default" in schema:
        return schema["default"]
    if schema.get("enum"):
        return schema["enum"][0]
    t = schema.get("type")
    if t == "array":
        return [_example_for_schema(schema.get("items") or {}, spec, depth + 1)]
    if t == "object" or "properties" in schema:
        result: dict[str, Any] = {}
        for name, prop in (schema.get("properties") or {}).items():
            result[name] = _example_for_schema(prop, spec, depth + 1)
        return result
    if t == "integer":
        return 0
    if t == "number":
        return 0.0
    if t == "boolean":
        return False
    fmt = schema.get("format")
    if fmt == "date-time":
        return "2026-01-01T00:00:00Z"
    if fmt == "uuid":
        return "00000000-0000-0000-0000-000000000000"
    if fmt == "email":
        return "user@example.com"
    return ""


def _flatten_schema(schema: dict[str, Any], spec: dict[str, Any]) -> list[dict[str, str]]:
    """Flatten a Pydantic-shaped object schema into one row per field."""
    if "$ref" in schema:
        schema = _resolve_ref(spec, schema["$ref"])
    if "allOf" in schema:
        merged: dict[str, Any] = {"type": "object", "properties": {}, "required": []}
        for piece in schema["allOf"]:
            piece = _resolve_ref(spec, piece["$ref"]) if "$ref" in piece else piece
            merged["properties"].update(piece.get("properties") or {})
            merged["required"].extend(piece.get("required") or [])
        schema = merged

    properties = schema.get("properties") or {}
    required = set(schema.get("required") or [])
    rows: list[dict[str, str]] = []
    for name, prop in properties.items():
        rows.append(
            {
                "field": name,
                "type": _schema_summary(prop, spec),
                "required": "yes" if name in required else "no",
                "default": json.dumps(prop["default"]) if "default" in prop else "",
                "description": prop.get("description") or prop.get("title") or "",
                "constraints": _constraints_text(prop),
            }
        )
    return rows


def _constraints_text(schema: dict[str, Any]) -> str:
    bits: list[str] = []
    for key in ("minLength", "maxLength", "minimum", "maximum", "pattern", "enum"):
        if key in schema and schema[key] is not None:
            bits.append(f"{key}={schema[key]}")
    if schema.get("nullable"):
        bits.append("nullable=true")
    return ", ".join(bits)


def _security_for_op(op: dict[str, Any], spec: dict[str, Any]) -> str:
    sec = op.get("security")
    if sec is None:
        # Operations inherit top-level security if any.
        sec = spec.get("security") or []
    if not sec:
        return "Public"
    names: list[str] = []
    for entry in sec:
        names.extend(entry.keys())
    return ", ".join(names) if names else "Bearer"


def _safe_sheet_name(name: str) -> str:
    invalid = set(r"[]:*?/\\")
    cleaned = "".join("-" if ch in invalid else ch for ch in name).strip()
    return cleaned[:31] or "Sheet"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_overview_sheet(wb: Workbook, spec: dict[str, Any]) -> None:
    ws = wb.active
    assert ws is not None
    ws.title = "Overview"

    info = spec.get("info") or {}
    style_hero(
        ws,
        title=f"{info.get('title', 'QualityForge AI')} — REST API Reference",
        subtitle=f"Version {info.get('version', '0.0.0')}  ·  generated from live OpenAPI",
        columns=2,
    )

    write_kv_block(
        ws,
        rows=[
            ("Service", info.get("title", "")),
            ("Version", info.get("version", "")),
            ("Description", info.get("description", "")),
            ("OpenAPI", spec.get("openapi", "")),
            ("Base URL (dev)", "http://localhost:8000/api/v1"),
            ("Auth", "Bearer JWT (Authorization: Bearer <access_token>)"),
            ("Content-Type", "application/json (multipart/form-data for uploads)"),
            ("Pagination", "?page=1&page_size=20  → meta.{page,page_size,total,total_pages}"),
            ("Sorting", "?sort=field   |  ?sort=-field (desc)"),
            ("Idempotency", "Idempotency-Key header on POST creates"),
            ("Errors", "{ error: { code, message, details }, meta: { request_id } }"),
            ("Live Docs", "GET /api/docs (Swagger UI)  ·  GET /api/openapi.json"),
        ],
        start_row=4,
    )

    section_title(ws, row=18, text="How to use this workbook", span=2)
    write_kv_block(
        ws,
        rows=[
            ("Index", "Every endpoint, sortable. Use Excel filtering (header arrows)."),
            ("<Tag> sheets", "One per API group — auth, projects, executions, stm, …"),
            ("Schemas", "Every Pydantic model with full field reference."),
            ("Enums", "Every enum value the API accepts."),
            ("Regenerate", "make excel-docs    (also runs build_api_workbook.py)"),
        ],
        start_row=19,
    )


def _build_index_sheet(wb: Workbook, spec: dict[str, Any]) -> list[tuple[str, str, str, str]]:
    """Create the index sheet and return (tag, method, path, summary) rows."""
    ws = wb.create_sheet("Index")
    style_hero(
        ws,
        title="API Endpoint Index",
        subtitle="All endpoints in the platform. Filter / sort like any Excel table.",
        columns=6,
    )

    rows: list[tuple[str, str, str, str]] = []
    table_rows: list[list[object]] = []
    for path, methods in (spec.get("paths") or {}).items():
        for method, op in methods.items():
            if method.lower() not in {"get", "post", "put", "patch", "delete"}:
                continue
            tags = op.get("tags") or ["misc"]
            for tag in tags:
                rows.append((tag, method.upper(), path, op.get("summary") or op.get("operationId") or ""))

    rows.sort(key=lambda r: (r[0], r[2], r[1]))
    for tag, method, path, summary in rows:
        table_rows.append(
            [
                tag,
                method,
                path,
                summary,
                _safe_sheet_name(tag),
            ]
        )

    last_row = write_table(
        ws,
        headers=["Tag", "Method", "Path", "Summary", "Detail Sheet"],
        rows=table_rows,
        start_row=4,
        column_widths=[18, 10, 65, 60, 22],
        table_name="EndpointIndex",
    )

    # Hyperlink the "Detail Sheet" column to its sheet.
    for excel_row in range(5, last_row + 1):
        sheet_cell = ws.cell(row=excel_row, column=5)
        target_sheet = sheet_cell.value
        if target_sheet:
            sheet_cell.hyperlink = f"#'{target_sheet}'!A1"
            sheet_cell.style = "Hyperlink"
    return rows


def _build_tag_sheet(wb: Workbook, tag: str, ops: list[dict[str, Any]], spec: dict[str, Any]) -> None:
    ws = wb.create_sheet(_safe_sheet_name(tag))
    style_hero(
        ws,
        title=f"{tag.capitalize()} — Endpoints",
        subtitle=f"{len(ops)} endpoint(s). Each block lists params, request, response, examples.",
        columns=6,
    )

    row = 4
    for op_record in ops:
        row = _render_endpoint_block(ws, op_record, spec, row)
        row += 1  # spacer
    autosize(ws, headers=["A", "B", "C", "D", "E", "F"], max_width=80)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 60


def _render_endpoint_block(
    ws: Worksheet, op_record: dict[str, Any], spec: dict[str, Any], start_row: int
) -> int:
    op = op_record["operation"]
    method = op_record["method"]
    path = op_record["path"]

    section_title(ws, row=start_row, text=f"{method}  {path}", span=6)
    row = start_row + 1

    summary = op.get("summary") or op.get("operationId") or ""
    description = op.get("description") or ""
    write_kv_block(
        ws,
        rows=[
            ("Summary", summary),
            ("Description", description),
            ("Operation ID", op.get("operationId") or ""),
            ("Auth", _security_for_op(op, spec)),
        ],
        start_row=row,
    )
    row += 4

    # ---- Parameters ----
    params = op.get("parameters") or []
    param_rows: list[list[object]] = []
    for p in params:
        schema = p.get("schema") or {}
        param_rows.append(
            [
                p.get("name", ""),
                p.get("in", ""),
                _schema_summary(schema, spec),
                "yes" if p.get("required") else "no",
                json.dumps(schema["default"]) if "default" in schema else "",
                p.get("description") or "",
            ]
        )
    if param_rows:
        ws.cell(row=row, column=1, value="Parameters").font = _bold_font()
        row += 1
        row = write_table(
            ws,
            headers=["Name", "In", "Type", "Required", "Default", "Description"],
            rows=param_rows,
            start_row=row,
            column_widths=[22, 10, 22, 10, 16, 60],
        ) + 1

    # ---- Request body ----
    request_body = op.get("requestBody") or {}
    if request_body:
        ws.cell(row=row, column=1, value="Request Body").font = _bold_font()
        row += 1
        content = request_body.get("content") or {}
        for media_type, media in content.items():
            ws.cell(row=row, column=1, value=f"  Content-Type: {media_type}").font = _muted_font()
            row += 1
            schema = media.get("schema") or {}
            field_rows = _flatten_schema(schema, spec)
            if field_rows:
                row = write_table(
                    ws,
                    headers=["Field", "Type", "Required", "Default", "Description", "Constraints"],
                    rows=[
                        [
                            f["field"],
                            f["type"],
                            f["required"],
                            f["default"],
                            f["description"],
                            f["constraints"],
                        ]
                        for f in field_rows
                    ],
                    start_row=row,
                    column_widths=[22, 28, 10, 16, 50, 28],
                ) + 1
            example = _example_for_schema(schema, spec)
            if example is not None:
                ws.cell(row=row, column=1, value="  Example body").font = _muted_font()
                ws.cell(row=row, column=2, value=json.dumps(example, indent=2)).alignment = _wrap()
                ws.row_dimensions[row].height = max(60, 16 * (str(example).count("\n") + 4))
                row += 1

    # ---- Responses ----
    responses = op.get("responses") or {}
    if responses:
        ws.cell(row=row, column=1, value="Responses").font = _bold_font()
        row += 1
        for status_code, resp in responses.items():
            desc = resp.get("description") or ""
            ws.cell(row=row, column=1, value=f"  {status_code}").font = _bold_font()
            ws.cell(row=row, column=2, value=desc).alignment = _wrap()
            row += 1
            content = resp.get("content") or {}
            for media_type, media in content.items():
                ws.cell(row=row, column=1, value=f"    Content-Type: {media_type}").font = _muted_font()
                row += 1
                schema = media.get("schema") or {}
                field_rows = _flatten_schema(schema, spec)
                if field_rows:
                    row = write_table(
                        ws,
                        headers=["Field", "Type", "Required", "Default", "Description", "Constraints"],
                        rows=[
                            [
                                f["field"],
                                f["type"],
                                f["required"],
                                f["default"],
                                f["description"],
                                f["constraints"],
                            ]
                            for f in field_rows
                        ],
                        start_row=row,
                        column_widths=[22, 28, 10, 16, 50, 28],
                    ) + 1
                example = _example_for_schema(schema, spec)
                if example is not None:
                    ws.cell(row=row, column=1, value="    Example response").font = _muted_font()
                    ws.cell(row=row, column=2, value=json.dumps(example, indent=2)).alignment = _wrap()
                    ws.row_dimensions[row].height = max(60, 16 * (str(example).count("\n") + 4))
                    row += 1
    return row


def _build_schemas_sheet(wb: Workbook, spec: dict[str, Any]) -> None:
    ws = wb.create_sheet("Schemas")
    style_hero(
        ws,
        title="Schemas — Pydantic models",
        subtitle="Every request/response schema, one row per field.",
        columns=7,
    )
    schemas = ((spec.get("components") or {}).get("schemas") or {})
    rows: list[list[object]] = []
    for schema_name, schema in sorted(schemas.items()):
        if "enum" in schema:
            continue  # handled in Enums sheet
        for f in _flatten_schema(schema, spec):
            rows.append(
                [
                    schema_name,
                    f["field"],
                    f["type"],
                    f["required"],
                    f["default"],
                    f["description"],
                    f["constraints"],
                ]
            )
    write_table(
        ws,
        headers=["Schema", "Field", "Type", "Required", "Default", "Description", "Constraints"],
        rows=rows,
        start_row=4,
        column_widths=[28, 22, 30, 10, 18, 50, 28],
        table_name="SchemaFields",
    )


def _build_enums_sheet(wb: Workbook, spec: dict[str, Any]) -> None:
    ws = wb.create_sheet("Enums")
    style_hero(
        ws,
        title="Enums",
        subtitle="Every enum value the API accepts (RunStatus, ValidationType, …).",
        columns=3,
    )
    schemas = ((spec.get("components") or {}).get("schemas") or {})
    rows: list[list[object]] = []
    for schema_name, schema in sorted(schemas.items()):
        if "enum" not in schema:
            continue
        for value in schema["enum"]:
            rows.append([schema_name, value, schema.get("description") or schema.get("title") or ""])
    write_table(
        ws,
        headers=["Enum", "Value", "Notes"],
        rows=rows,
        start_row=4,
        column_widths=[28, 22, 60],
        table_name="EnumValues",
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _bold_font():
    from openpyxl.styles import Font

    return Font(name="Calibri", size=11, bold=True, color="FF1E1B4B")


def _muted_font():
    from openpyxl.styles import Font

    return Font(name="Calibri", size=10, italic=True, color="FF475569")


def _wrap():
    from openpyxl.styles import Alignment

    return Alignment(vertical="top", horizontal="left", wrap_text=True)


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------


def build(output: Path) -> Path:
    spec = _load_openapi()

    wb = Workbook()
    _build_overview_sheet(wb, spec)
    _build_index_sheet(wb, spec)

    # Bucket operations by primary tag.
    by_tag: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for path, methods in (spec.get("paths") or {}).items():
        for method, op in methods.items():
            if method.lower() not in {"get", "post", "put", "patch", "delete"}:
                continue
            tag = (op.get("tags") or ["misc"])[0]
            by_tag[tag].append({"path": path, "method": method.upper(), "operation": op})

    for tag in sorted(by_tag):
        by_tag[tag].sort(key=lambda r: (r["path"], r["method"]))
        _build_tag_sheet(wb, tag, by_tag[tag], spec)

    _build_schemas_sheet(wb, spec)
    _build_enums_sheet(wb, spec)

    configure_workbook(wb)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "docs" / "excel" / "qualityforge-ai_api_reference.xlsx",
    )
    args = parser.parse_args()
    path = build(args.out)
    # Use plain print: this script is run from the CLI / Makefile.
    print(f"✓ wrote {path}")


if __name__ == "__main__":
    main()
