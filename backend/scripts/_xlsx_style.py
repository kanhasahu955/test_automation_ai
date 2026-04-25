"""Shared styling helpers used by the Excel-doc generators.

Keeps every workbook visually consistent (header bar, freeze pane, column
widths, borders) without duplicating the pattern across generators.
"""
from __future__ import annotations

from collections.abc import Iterable, Sequence

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Brand palette — kept in sync with frontend/src/theme/tokens.ts.
# ---------------------------------------------------------------------------
COLOR_BRAND_PRIMARY = "FF6366F1"   # indigo-500
COLOR_BRAND_DARK = "FF1E1B4B"      # indigo-950 hero band
COLOR_BRAND_ACCENT = "FF06B6D4"    # cyan-500
COLOR_BRAND_SOFT = "FFEEF2FF"      # indigo-50
COLOR_GREY_50 = "FFF8FAFC"
COLOR_GREY_100 = "FFF1F5F9"
COLOR_GREY_200 = "FFE2E8F0"
COLOR_TEXT_DARK = "FF0F172A"
COLOR_TEXT_MUTED = "FF475569"


def _thin_border(color: str = COLOR_GREY_200) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


THIN_BORDER = _thin_border()


def style_hero(ws: Worksheet, *, title: str, subtitle: str, columns: int) -> None:
    """Render a two-row hero band (title + subtitle) at the top of a sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)

    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=18, bold=True, color="FFFFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_BRAND_DARK)
    title_cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30

    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color="FFCBD5E1")
    sub_cell.fill = PatternFill("solid", fgColor=COLOR_BRAND_DARK)
    sub_cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 22


def style_header_row(ws: Worksheet, row: int, columns: int) -> None:
    """Style a single header row with the brand-primary fill + white bold text."""
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLOR_BRAND_PRIMARY)
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def write_table(
    ws: Worksheet,
    *,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    start_row: int,
    column_widths: Sequence[int] | None = None,
    zebra: bool = True,
    table_name: str | None = None,
) -> int:
    """Write headers + rows starting at `start_row`. Returns the last row written."""
    for col_idx, name in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=name)
    style_header_row(ws, start_row, len(headers))

    last_row = start_row
    for r_idx, row in enumerate(rows, start=1):
        excel_row = start_row + r_idx
        last_row = excel_row
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c_idx, value=_coerce(value))
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            cell.font = Font(name="Calibri", size=10, color=COLOR_TEXT_DARK)
            cell.border = THIN_BORDER
            if zebra and r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLOR_GREY_50)

    if column_widths:
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    else:
        autosize(ws, headers=headers, max_width=60)

    if table_name and last_row > start_row:
        ref = f"A{start_row}:{get_column_letter(len(headers))}{last_row}"
        tbl = Table(displayName=_safe_table_name(table_name), ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # We've applied custom styles already; an Excel "table" gives users
        # filtering/sorting by default. Wrapping it adds those affordances.
        ws.add_table(tbl)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return last_row


def autosize(ws: Worksheet, headers: Sequence[str], *, max_width: int = 60) -> None:
    """Approximate column auto-size based on the longest cell value."""
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            value = cell.value
            if value is None:
                continue
            for line in str(value).splitlines() or [""]:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), max_width)


def write_kv_block(
    ws: Worksheet,
    *,
    rows: Sequence[tuple[str, object]],
    start_row: int,
    label_width: int = 22,
    value_width: int = 80,
) -> int:
    """Write a vertical key/value block (label column + value column)."""
    for offset, (label, value) in enumerate(rows):
        row = start_row + offset
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_MUTED)
        label_cell.alignment = Alignment(vertical="top", horizontal="left", indent=1)
        label_cell.fill = PatternFill("solid", fgColor=COLOR_BRAND_SOFT)
        label_cell.border = THIN_BORDER

        val_cell = ws.cell(row=row, column=2, value=_coerce(value))
        val_cell.font = Font(name="Calibri", size=11, color=COLOR_TEXT_DARK)
        val_cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        val_cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = label_width
    ws.column_dimensions["B"].width = value_width
    return start_row + len(rows) - 1


def section_title(ws: Worksheet, *, row: int, text: str, span: int = 6) -> None:
    """Render a colored section subheading."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=COLOR_BRAND_ACCENT)
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 22


def configure_workbook(wb: Workbook) -> None:
    """Workbook-level defaults — view zoom, default font, etc."""
    if wb.active is not None:
        wb.active.sheet_view.showGridLines = False
    for sheet in wb.worksheets:
        sheet.sheet_view.zoomScale = 110
        sheet.sheet_view.showGridLines = False


def _coerce(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _safe_table_name(name: str) -> str:
    """Excel table names: letters/digits/underscores, must start with a letter."""
    safe = "".join(c if c.isalnum() else "_" for c in name)
    if not safe or not safe[0].isalpha():
        safe = "T_" + safe
    return safe[:255]
